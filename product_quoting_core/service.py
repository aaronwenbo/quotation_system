#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
统一报价服务 - Web应用和CLI技能共享此服务
"""

import logging
import pandas as pd
from pathlib import Path
from typing import Tuple, List, Dict, Optional
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from .library import StandardLibrary, MatchingRules
from .matcher import CodeMatcher

logger = logging.getLogger(__name__)


class QuotingService:
    """统一报价服务 - 完整的5条规则实现"""

    def __init__(self, lib_path: str, rules_path: str, log_dir: Optional[str] = None):
        self.lib = StandardLibrary(lib_path)
        self.rules = MatchingRules(rules_path)
        self.matcher = CodeMatcher(self.lib, self.rules)
        self.log_dir = Path(log_dir) if log_dir else None
        # 更新日志文件路径
        if self.log_dir:
            self.update_log_path = self.log_dir / 'library_updates.xlsx'
        else:
            self.update_log_path = None

    def _write_update_log(self, added: List[Dict]) -> None:
        """
        写入标准库更新日志

        Args:
            added: 新增产品列表，每个包含 original, std, price
        """
        if not self.update_log_path or not added:
            return

        # 确保日志目录存在
        self.update_log_path.parent.mkdir(parents=True, exist_ok=True)

        now = datetime.now()
        timestamp = now.strftime('%Y-%m-%d %H:%M:%S')

        # 构建日志数据
        log_entries = []
        for item in added:
            log_entries.append({
                '加入时间': timestamp,
                '标准编码': item['std'],
                '原始编码': item['original'],
                '价格': float(item['price'])
            })

        # 读取现有日志（如果存在）
        if self.update_log_path.exists():
            existing_df = pd.read_excel(self.update_log_path)
            new_df = pd.DataFrame(log_entries)
            combined_df = pd.concat([new_df, existing_df], ignore_index=True)
        else:
            combined_df = pd.DataFrame(log_entries)

        # 保存日志
        combined_df.to_excel(self.update_log_path, index=False)
        logger.info(f"更新日志已写入: {len(added)} 条记录")

    def get_update_log(self, limit: int = 100) -> List[Dict]:
        """
        获取标准库更新日志

        Args:
            limit: 返回最近的记录数

        Returns:
            更新日志列表
        """
        if not self.update_log_path or not self.update_log_path.exists():
            return []

        df = pd.read_excel(self.update_log_path)
        # 转换为字典列表，最多返回limit条
        return df.head(limit).to_dict('records')

    def _parse_quantity(self, raw_value) -> Tuple[Optional[float], Optional[str]]:
        """
        将数量列的值转换为数字

        支持格式：纯数字、千分位逗号、中文单位（个/台/套/只/件/支/米/卷）、空格
        若无法解析则返回 None 和警告信息

        Returns:
            (数字值或None, 警告信息或None)
        """
        if raw_value is None:
            return None, None

        # 转为字符串处理
        s = str(raw_value).strip()
        if not s or s.lower() == 'nan':
            return None, None

        # 记录原始值用于警告
        original = s

        # 去掉常见中文单位
        units = ['个', '台', '套', '只', '件', '支', '米', '卷', '箱', '包', '桶', '瓶', '张', '根', '对', '付', '条', '片', '块', '把']
        for unit in units:
            if s.endswith(unit):
                s = s[:-1].strip()
                break

        # 去掉千分位逗号（1,234 → 1234）
        s = s.replace(',', '')

        try:
            return float(s), None
        except (ValueError, TypeError):
            return None, f"第{0}行数量无法识别为数字: '{original}'"  # 行号由调用方填充

    def process_quote(self, file_path: str, code_col: int, qty_col: int) -> Tuple[pd.DataFrame, Dict, List[str]]:
        """
        处理订单报价

        Args:
            file_path: 订单文件路径
            code_col: 产品编码列号 (A=0, B=1...)
            qty_col: 数量列号

        Returns:
            (结果DataFrame, 统计字典, 未匹配编码列表)
        """
        if str(file_path).endswith('.csv'):
            df = pd.read_csv(file_path, header=None)
        else:
            df = pd.read_excel(file_path, header=None)

        stats = {
            'total': 0,
            'direct_match': 0,
            'rule1_match': 0,
            'rule2_match': 0,
            'rule3_match': 0,
            'rule4_match': 0,
            'rule5_match': 0,
            'rule_combination': 0,
            'no_match': 0
        }
        no_match_codes = []
        qty_warnings = []  # 数量列无法解析的警告列表

        for idx, row in df.iterrows():
            code = row.iloc[code_col] if code_col < len(row) else None
            if pd.isna(code) or not str(code).strip():
                continue

            stats['total'] += 1
            matched_code, product_info, label = self.matcher.match(code)

            price_col_idx = max(code_col, qty_col) + 1
            total_price_col_idx = price_col_idx + 1
            std_code_col_idx = price_col_idx + 2
            orig_code_col_idx = price_col_idx + 3
            label_col_idx = price_col_idx + 4

            # 确保列存在
            for col_idx in [price_col_idx, total_price_col_idx, std_code_col_idx, orig_code_col_idx, label_col_idx]:
                while len(df.columns) <= col_idx:
                    df[len(df.columns)] = None

            if product_info:
                df.iloc[idx, price_col_idx] = product_info['价格']
                qty_raw = row.iloc[qty_col] if qty_col < len(row) else None
                if pd.notna(qty_raw):
                    qty, warning = self._parse_quantity(qty_raw)
                    if qty is not None:
                        df.iloc[idx, total_price_col_idx] = qty * float(product_info['价格'])
                    if warning:
                        # 填充行号
                        excel_row = idx + 1  # DataFrame 0-based → Excel 1-based
                        warning_with_row = warning.replace("第0行", f"第{excel_row}行")
                        qty_warnings.append(warning_with_row)
                        logger.warning(warning_with_row)
                df.iloc[idx, std_code_col_idx] = matched_code
                df.iloc[idx, orig_code_col_idx] = product_info.get('原规格编码', '')
                df.iloc[idx, label_col_idx] = label

                # 统计匹配类型
                if label == '直接匹配':
                    stats['direct_match'] += 1
                elif '规则一' in label:
                    stats['rule1_match'] += 1
                elif '规则五' in label:
                    stats['rule5_match'] += 1
                elif '规则四' in label:
                    stats['rule4_match'] += 1
                elif '去T' in label and '去*' not in label:
                    stats['rule2_match'] += 1
                elif '去*' in label and '去T' not in label:
                    stats['rule3_match'] += 1
                elif '去T' in label and '去*' in label:
                    stats['rule_combination'] += 1
                else:
                    stats['rule_combination'] += 1
            else:
                df.iloc[idx, label_col_idx] = label
                stats['no_match'] += 1
                no_match_codes.append(str(code).strip())

        stats['qty_warnings'] = qty_warnings
        return df, stats, no_match_codes

    def update_library(self, file_path: str, code_col: int, price_col: int, label_col: int) -> Tuple[List[Dict], List[Dict]]:
        """
        从用户上传文件更新标准库

        Args:
            file_path: 反馈文件路径
            code_col: 产品编码列号
            price_col: 价格列号
            label_col: 匹配标注列号

        Returns:
            (新增列表, 跳过列表)
        """
        if str(file_path).endswith('.csv'):
            df = pd.read_csv(file_path, header=None)
        else:
            df = pd.read_excel(file_path, header=None)

        added = []
        skipped = []

        for idx, row in df.iterrows():
            if idx == 0:  # 跳过标题行
                continue

            code = row.iloc[code_col] if code_col < len(row) else None
            price = row.iloc[price_col] if price_col < len(row) else None
            label = row.iloc[label_col] if label_col < len(row) else None

            # 只处理无匹配且有价格的
            if label != '无匹配':
                continue
            if pd.isna(price):
                continue
            if pd.isna(code) or not str(code).strip():
                continue

            original_code = self.matcher.clean_code(code)
            std_code, _ = self.matcher.clean_for_library(original_code)

            if self.lib.has(std_code):
                skipped.append({'original': original_code, 'std': std_code})
            else:
                try:
                    price_value = float(price)
                    self.lib.add(std_code, price_value, '', original_code)
                    added.append({'original': original_code, 'std': std_code, 'price': price_value})
                except (ValueError, TypeError):
                    skipped.append({'original': original_code, 'std': std_code, 'reason': '价格无效'})

        if added:
            self.lib.save()
            # 写入更新日志
            self._write_update_log(added)

        return added, skipped

    def backup_library(self, backup_dir: str) -> Path:
        """备份标准库"""
        return self.lib.backup(backup_dir)

    def get_library_size(self) -> int:
        """获取标准库大小"""
        return self.lib.size()

    def check_duplicates(self) -> List[str]:
        """检查重复编码"""
        return self.lib.check_duplicates()

    def save_styled_quote(self, df: pd.DataFrame, output_path: str, code_col: int, qty_col: int) -> None:
        """
        保存带样式的报价结果Excel

        样式说明：
        - 表头重命名：单价、总价、标准编码、原规格编码、匹配标注
        - 列宽自适应调整
        - 匹配标注列颜色：
          * 直接匹配 → 绿色背景
          * 规则匹配 → 黄色背景
          * 无匹配 → 红色背景

        Args:
            df: 处理后的DataFrame
            output_path: 输出文件路径
            code_col: 产品编码列号
            qty_col: 数量列号
        """
        # 计算列位置
        price_col_idx = max(code_col, qty_col) + 1
        label_col_idx = price_col_idx + 4

        # 先保存为普通Excel（为了处理中文字符）
        df.to_excel(output_path, index=False, header=False)

        # 用openpyxl重新加载并应用样式
        wb = load_workbook(output_path)
        ws = wb.active

        # ========== 1. 设置表头（第一行） ==========
        headers = {
            price_col_idx: '单价',
            price_col_idx + 1: '总价',
            price_col_idx + 2: '标准编码',
            price_col_idx + 3: '原规格编码',
            price_col_idx + 4: '匹配标注'
        }
        for col_idx, header_text in headers.items():
            cell = ws.cell(row=1, column=col_idx + 1)  # openpyxl是1-based
            cell.value = header_text
            cell.font = Font(bold=True)

        # ========== 2. 调整列宽 ==========
        # 计算每列的最大宽度
        column_widths = {}
        for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row)):
            for cell in row:
                if cell.value:
                    col_letter = get_column_letter(cell.column)
                    # 计算中文字符的宽度（每个中文字符约2个英文字符宽度）
                    text = str(cell.value)
                    chinese_count = sum(1 for c in text if '一' <= c <= '鿿')
                    width = len(text) + chinese_count + 4  # 额外留边距
                    if col_letter not in column_widths or width > column_widths[col_letter]:
                        column_widths[col_letter] = width

        # 应用列宽（设置最大限制为50）
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = min(width, 50)

        # ========== 3. 匹配标注列背景色设置 ==========
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 浅绿
        yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # 浅黄
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')     # 浅红

        label_col_excel = label_col_idx + 1  # 转换为openpyxl的1-based

        for row_idx in range(2, ws.max_row + 1):  # 从第2行开始（跳过表头）
            cell = ws.cell(row=row_idx, column=label_col_excel)
            label_value = str(cell.value).strip() if cell.value else ''

            if not label_value:
                continue
            elif label_value == '直接匹配':
                cell.fill = green_fill
            elif label_value == '无匹配':
                cell.fill = red_fill
            elif '规则' in label_value or '去T' in label_value or '去*' in label_value:
                cell.fill = yellow_fill

        # 保存文件
        wb.save(output_path)
        logger.info(f"带样式的报价结果已保存: {output_path}")
